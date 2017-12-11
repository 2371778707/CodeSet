#!/usr/bin/env python3
#-*-coding:utf-8-*-
# __all__=""
# __datetime__=""
# __purpose__=""

#!/usr/bin/env python3
#-*-coding:utf-8-*-
# __all__=""
# __datetime__=""
# __purpose__="整理新闻导入excel格式为"area","url""

from foreigin_all_news.KTspider.rules import Rules
rule_list = Rules.rules_list


def convert(rule, excel_name, title=None):
    """
    :param rule: 规则
    :param excel_name:表名
    :param title: 表分类名
    :return:excel
    """
    # 控制表头
    if not title:
        title = "默认"
    list1 = []
    # 格式化数据
    list1.append([rule[0].get("area"), rule[0].get("url")])
    for i in range(1, len(rule) - 1):
        if rule[i + 1].get("area") != rule[i].get("area"):
            list1.append([rule[i + 1].get("area"), rule[i + 1].get("url")])
        else:
            for j in list1:
                if rule[i].get("area") == j[0]:
                    for ii in rule[i + 1].get("url"):
                        j[1].append(ii)
    import openpyxl
    # 新建一个excel表
    wb = openpyxl.Workbook()
    wb.save(excel_name)
    wb = openpyxl.load_workbook(excel_name)
    sheet = wb.active
    sheet.title = title
    # 设置表头
    name_str = ["", "area", "url", ""]
    for field in range(1, 3):
        _ = sheet.cell(row=1, column=field, value=name_str[field])
    # 写入数据
    for i in range(0, len(list1)):
        _ = sheet.cell(row=i + 2, column=1, value=list1[i][0])
    for i in range(0, len(list1)):
        _ = sheet.cell(row=i + 2, column=2, value=','.join(list1[i][1]))
    wb.save(excel_name)


convert(rule_list, "a.xlsx")